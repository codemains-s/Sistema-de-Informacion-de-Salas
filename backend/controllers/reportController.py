from sqlalchemy.orm import Session
from models.tables import User, CompletedHours, RoomBooking
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import requests
from PIL import Image as PILImage


def generate_user_report(user_id: int, db: Session):
    """
    Genera un reporte en Excel para un usuario específico con sus reservas de sala y horas completadas.

    Args:
        user_id (int): ID del usuario para generar el reporte
        db (Session): Sesión de la base de datos

    Returns:
        BytesIO: Un objeto BytesIO que contiene el archivo Excel generado, o None si el usuario no existe

    El reporte incluye:
        - Fecha de reserva
        - Nombre de la sala
        - Hora de inicio y fin
        - Observaciones
        - Nombre del usuario
        - Horas acumuladas
        - Firma del coordinador (como imagen)

    La función realiza las siguientes operaciones:
        1. Obtiene los datos del usuario
        2. Recupera las reservas de sala y horas completadas
        3. Genera un DataFrame con la información
        4. Crea un archivo Excel con formato personalizado
        5. Añade las firmas como imágenes en el archivo
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        return None

    room_bookings = db.query(RoomBooking).filter(RoomBooking.user_id == user_id).all()
    completed_hours = (
        db.query(CompletedHours).filter(CompletedHours.user_id == user_id).all()
    )

    report_data = []
    accumulated_hours = 0
    for booking, hour in zip(room_bookings, completed_hours):
        completed = hour.hour_completed if hour else 0
        accumulated_hours += completed

        booking_date = booking.booking_date if booking else pd.NaT

        report_data.append(
            {
                "Fecha": booking_date,
                "Sala": booking.room.name if booking else pd.NA,
                "Hora Inicio": booking.start_time if booking else pd.NA,
                "Hora Fin": booking.end_time if booking else pd.NA,
                "Observaciones": hour.observations if hour else pd.NA,
                "Nombre": user.name,
                "Horas Completadas": accumulated_hours,
                "Coordinador": None,
            }
        )

    report_df = pd.DataFrame(report_data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="User Report")

        workbook = writer.book
        worksheet = writer.sheets["User Report"]

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Añadir la imagen de la firma
        for i, hour in enumerate(completed_hours, start=2):
            signature_url = hour.signature
            if signature_url:
                try:
                    # Obtener la imagen desde la URL y redimensionarla
                    response = requests.get(signature_url)
                    img = PILImage.open(BytesIO(response.content))
                    img = img.resize((60, 40), PILImage.LANCZOS)
                    img_byte_arr = BytesIO()
                    img.save(img_byte_arr, format="PNG")
                    img_byte_arr.seek(0)

                    excel_img = Image(img_byte_arr)
                    excel_img.anchor = f"H{str(i)}"

                    worksheet.add_image(excel_img)

                    worksheet.row_dimensions[i].height = 40

                except Exception as e:
                    print(f"Error al agregar la firma: {e}")

        for col_idx, col_cells in enumerate(worksheet.columns, start=1):
            max_length = (
                max(len(str(cell.value)) for cell in col_cells if cell.value) + 2
            )
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = max_length

    output.seek(0)
    return output
